import pdfplumber
import re
import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from gooey import Gooey, GooeyParser


def re_text(bt, text):
    m1 = re.search(bt, text)
    if m1 is not None:
        return re_block(m1[0])


def re_block(text):
    return text.replace(' ', '').replace('　', '').replace('）', '').replace(')', '').replace('：', ':')


def get_all(dir_path):
    all_files_2 = []
    all_names_2 = []
    folder_path = []

    for root, sub_dirs, file_names in os.walk(dir_path):
        all_files = []
        all_names = []
        root = root.replace('\\', '/')
        for name in file_names:
            all_path = os.path.join(root, name)
            all_path = all_path.replace('\\', '/')
            all_files.append(all_path)
            all_names.append(name)
        all_files_2.append(all_files)
        all_names_2.append(all_names)
        folder_path.append(root)

    return all_files_2, all_names_2, folder_path


def read(root_dir):
    all_files_2, all_names_2, folder_paths = get_all(root_dir)
    return all_files_2, all_names_2, folder_paths


def new_folder(i, root_dir, out_dir, old_dir):
    new_dir = old_dir.replace(root_dir, out_dir)
    if i != 0:
        new_sheet = old_dir.replace(root_dir + "/", "")
        new_sheet = new_sheet.replace("/", "_")
    else:
        new_sheet = old_dir.split("/")[-1]
    if not os.path.exists(new_dir):
        # 如果目标路径不存在原文件夹的话就创建
        os.makedirs(new_dir)

    if os.path.exists(new_dir):
        # 如果目标路径存在原文件夹的话就先删除
        shutil.rmtree(new_dir)
        os.makedirs(new_dir)

    return new_dir, new_sheet


def check_name(pdf_text, invoice_name=None, tax_id=None):

    if "91440300MA5GC316X2" in pdf_text or "9 1440300MA5GC316X2" in pdf_text:
        tax_id = "对了"
        if "佛山市顺德区瑞磐科技有限公司龙华分公司" in pdf_text:
            invoice_name = "对了"
        else:
            invoice_name = "×了"
    else:
        tax_id = "×了"
        if "佛山市顺德区瑞磐科技有限公司龙华分公司" in pdf_text:
            invoice_name = "对了"

        else:
            invoice_name = "×了"
    return invoice_name, tax_id


def save_pdf(filepaths, filenames, folderpaths, out_dir, root_dir):

    for i, filepath in enumerate(filepaths):
        Repeat_name_list = []
        Repeat_num_list = []
        new_out_dir, new_sheet = new_folder(
            i, root_dir, out_dir, folderpaths[i])
        if i == 0:
            new_root_dir = new_out_dir
        for j, filepathNeed in enumerate(filepath):
            if filenames[i][j].endswith('.pdf') or filenames[i][j].endswith('.PDF'):
                tax_id = "没检查"
                invoice_name = "没检查"
                with pdfplumber.open(filepathNeed) as pdf:
                    first_page = pdf.pages[0]
                    pdf_text = first_page.extract_text()
                    invoice_name, tax_id = check_name(
                        pdf_text, invoice_name, tax_id)
                    try:
                        list_excel = re_info_1(
                            pdf_text, tax_id, invoice_name, filenames[i][j])
                        out_file_name = list_excel[4] + \
                            "-" + str(list_excel[5])
                    except:
                        list_excel = re_info_2(filenames[i][j])
                        out_file_name = list_excel[4]

                list_excel, Repeat_num_list = find_repeat_num(
                    list_excel, Repeat_num_list)

                list_excel, Repeat_name_list = find_repeat_name(
                    list_excel, Repeat_name_list, out_file_name)

                # copy_rename(i, filepathNeed, new_out_dir,new_root_dir, new_sheet, list_excel)
                copy_rename(filepathNeed, new_out_dir,
                            new_root_dir, new_sheet, list_excel)
            else:
                new_out = filepathNeed.replace(root_dir, new_root_dir)
                shutil.copy(filepathNeed, new_out)

    return "OK"


def find_repeat_name(list_excel, Repeat_name_list, out_file_name):
    Repeat_name_list.append(out_file_name)
    if out_file_name in Repeat_name_list:
        repeat_num = Repeat_name_list.count(out_file_name)
        if repeat_num == 1:
            out_file_name = out_file_name
        elif list_excel[3] == "重复":
            out_file_name = out_file_name + "(重复)"
        else:
            out_file_name = out_file_name + "_" + str(repeat_num)

        list_excel[7] = out_file_name + ".pdf"
        list_excel[7] = list_excel[7].replace(".0.", ".00.").replace(".0_", ".00_").replace(".0(", ".00(")

    return list_excel, Repeat_name_list


def find_repeat_num(list_excel, Repeat_num_list):
    if list_excel is not None:
        Repeat_num_list.append(list_excel[2])
        if list_excel[2] in Repeat_num_list:
            repeat_num = Repeat_num_list.count(list_excel[2])
            if repeat_num > 1:
                list_excel[3] = "重复"

    return list_excel, Repeat_num_list


def copy_rename(filepath, out_dir, new_root_dir, new_sheet, list_excel):

    dst = os.path.join(out_dir, list_excel[7])

    shutil.copy(filepath, dst)

    save_excel_2(new_root_dir, sheet_name=new_sheet, value_list=list_excel)

    return None


def re_info_1(pdf_text, tax_id, invoice_name, filename):
    list_excel = []

    # invoice_name = (
    #     re_text(re.compile(r'名\s*称\s*[:：]\s*([\u4e00-\u9fa5]+)'), pdf_text))
    # invoice_name = invoice_name.split(":", 1)[-1]

    invoice_number = (re_text(re.compile(r'发票号码(.*\d+)'), pdf_text))
    invoice_number = invoice_number.split(":", 1)[-1]

    invoice_date = (re_text(re.compile(r'开票日期(.*)'), pdf_text))
    invoice_date = invoice_date.split(":", 1)[-1]

    invoice_type_true = (
        re_text(re.compile(r'([/*]+[\u4e00-\u9fa5]+[ ])'), pdf_text))
    if invoice_type_true is None:
        invoice_type_true = (
            re_text(re.compile(r'([/（]+[\u4e00-\u9fa5]+[/）])'), pdf_text))
        invoice_type_true = invoice_type_true.split("（", 1)[-1]
    invoice_type_true = invoice_type_true.split("*", 1)[-1]

    if "服务费" in pdf_text and "电费" in pdf_text and "住宿" not in pdf_text:
        invoice_type_true = "充电费"
    elif "服务费" in pdf_text and "住宿" in pdf_text:
        invoice_type_true = "住宿费"

    # tax_id = (re_text(re.compile(r'纳税人识别号\s*[:：]\s*([a-zA-Z0-9]+)'), pdf_text))
    # tax_id = tax_id.split(":", 1)[-1]

    total_price_ture = re_text(re.compile(r'(小写.*(.*[0-9.]+))'), pdf_text)
    if total_price_ture is None:
        total_price_ture = re.findall(
            (r'(￥.*[0-9.]+)|(¥.*[0-9.]+)'), pdf_text)[-1]
        total_price_ture = ''.join(total_price_ture)
    if "￥" in total_price_ture:
        total_price_ture = total_price_ture.split("￥", 1)[-1]
    elif "¥" in total_price_ture:
        total_price_ture = total_price_ture.split("¥", 1)[-1]

    license_plate= re_text(re.compile(r'粤[a-zA-Z0-9]{6}'), pdf_text)

    list_excel.append(invoice_name)
    list_excel.append(tax_id)
    list_excel.append(int(invoice_number))
    list_excel.append("无重复")
    list_excel.append(invoice_type_true)
    list_excel.append(float(total_price_ture))
    list_excel.append(invoice_date)
    list_excel.append("")
    list_excel.append(filename)
    list_excel.append(license_plate)

    return list_excel


def re_info_2(filename):
    list_excel = []

    list_excel.append("")
    list_excel.append("")
    list_excel.append("")
    list_excel.append("")
    list_excel.append("其他单或读不出来")
    list_excel.append("")
    list_excel.append("")
    list_excel.append("")
    list_excel.append(filename)

    return list_excel


def save_excel_2(path, sheet_name, value_list=[[]]):
    # def save_excel_2(i, path, sheet_name, value_list=[[]]):
    value_list = [value_list]
    path = path + "/A_DEMO_汇总统计.xlsx"
    try:
        workbook = openpyxl.load_workbook(path)
        Sheet = workbook[sheet_name]
    except:
        try:
            workbook = openpyxl.load_workbook(path)
        except:
            workbook = openpyxl.Workbook()
        # workbook.create_sheet(sheet_name, i+1)
        workbook.create_sheet(sheet_name)
        Sheet = workbook[sheet_name]
        title = ['抬头', '纳税号', '发票号码', '发票',
                 '发票类型', '金额汇总', '开票日期', '改后pdf名称', '原pdf名称','车牌']
        Sheet.append(title)
        workbook.save(path)
    it_exist = find_exist(Sheet, value_list)
    if not it_exist:
        Sheet.append(value_list[0])
        for item, value in enumerate(value_list[0]):
            Sheet.column_dimensions[get_column_letter(
                item+1)].width = 1.8*len(str(value)) + 3
        workbook.save(path)

    workbook.close()
    return "ok"


def find_exist(Sheet, value_list):
    for row in Sheet.rows:
        row_data = [cell.value if cell.value != None else "" for cell in row]
        if value_list[0] == row_data:
            return True
    return False


def rename_pdf(filepaths, filenames):
    for i, filepath in enumerate(filepaths):
        Rename_list = []
        for j, filepathNeed in enumerate(filepath):
            if filenames[i][j].endswith('.pdf') or filenames[i][j].endswith('.PDF'):
                try:
                    filenameNeed = filenames[i][j].split("-")[-1]
                    if "_" in filenameNeed:
                        filenameNeed = filenameNeed.split("_")[0] + ".pdf"
                    Rename_list.append(filenameNeed)
                    new_name = filepathNeed.replace(
                        filenames[i][j], filenameNeed)
                    if filenameNeed in Rename_list:
                        repeat_num = Rename_list.count(filenameNeed)
                        if str(repeat_num) != "1":
                            new_name = new_name.replace(".00", "").replace(".PDF", "").replace(
                                ".pdf", "") + "-" + str(repeat_num) + ".pdf"
                        else:
                            new_name = new_name.replace(".00", "").replace(".PDF", "").replace(
                                ".pdf", "") + ".pdf"
                    os.rename(filepathNeed, new_name)
                except:
                    pass
    return "ok"


@Gooey(program_name="发票小工具", language='chinese', clear_before_run=True,
       encoding="utf-8", progress_regex=r"^progress: (\d+)%$", navigation='TABBED')
def GUI():
    know_text = "注意：使用时候不要打开对应的文件夹\n1.PDF复制后重新命名 2.发票信息保存至EXCEL 3.文件夹里面的重名文件"
    parser = GooeyParser(description=know_text)

    subs = parser.add_subparsers(
        help='commands', dest='command', required=False)

    my_parser = subs.add_parser('初始化pdf并生成Excel')
    group_1 = my_parser.add_argument_group(
        "(￣﹁￣)", gooey_options={'show_border': False, 'columns': 1})
    group_1.add_argument(
        '1.输入路径', help="原PDF文件的所在文件夹", widget="DirChooser")

    siege_parser = subs.add_parser('简单格式化pdf的文件名')
    group_2 = siege_parser.add_argument_group(
        "╮(￣▽ ￣)╭", gooey_options={'show_border': False, 'columns': 1})
    group_2.add_argument(
        '2.输入路径', help='要重命名文件的所在文件夹', widget="DirChooser")

    args = parser.parse_args()
    root_dir = vars(args).get("1.输入路径")
    rename_dir = vars(args).get("2.输入路径")

    return root_dir, rename_dir


if __name__ == '__main__':
    root_dir, rename_dir = GUI()

    if root_dir is not None and root_dir != ' ':
        out_dir = root_dir + "_输出"
        root_dir = root_dir.replace("\\", '/').replace("//", '/')
        out_dir = out_dir.replace("\\", '/').replace("//", '/')
        filepaths, filenames, folderpaths = read(root_dir.strip())
        if filepaths != []:
            save_pdf(filepaths, filenames, folderpaths,
                     out_dir.strip(), root_dir.strip())
            print("0、请使用pdf原文件")
            print("1、成功复制PDF后重新命名")
            print("2、成功保存发票信息至EXCEL")
        else:
            print("生成Excel的时候，请选择正确的文件夹!")
    elif rename_dir is not None and rename_dir != ' ':
        rename_dir = rename_dir.replace("\\", '/').replace("//", '/')
        filepaths, filenames, folderpaths = read(rename_dir.strip())
        if filepaths != []:
            rename_pdf(filepaths, filenames)
            print("0、请使用pdf原文件")
            print("1、成功改名保持并替代元文件")
        else:
            print("重新命名文件的时候，请选择正确的文件夹!")
