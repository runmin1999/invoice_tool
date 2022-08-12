import pathlib

from queue import Queue
from threading import Thread


import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import *
from tkinter.filedialog import askdirectory

import windnd

import pdfplumber
import re
import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

import time


class FileSearchEngine(ttk.Frame):

    # queue = Queue()
    # searching = False

    def __init__(self, master):
        super().__init__(master, padding=15)
        self.pack(fill=BOTH, expand=YES)

        # application variables
        _path = pathlib.Path().absolute().as_posix()
        self.path_var = ttk.StringVar(value=_path)
        self.type_var = ttk.StringVar(value='contains')

        # header and labelframe option container
        option_text = "(￣﹁￣)"
        self.option_lf = ttk.Labelframe(self, text=option_text, padding=15)
        self.option_lf.pack(fill=X, expand=YES, anchor=N)

        self.create_path_row()
        self.create_type_row()

        self.progressbar = ttk.Progressbar(
            master=self, 
            mode=INDETERMINATE, 
            bootstyle=(STRIPED, SUCCESS)
        )
        self.progressbar.pack(fill=X, expand=YES)


    def create_path_row(self):
        """Add path row to labelframe"""
        path_row = ttk.Frame(self.option_lf)
        path_row.pack(fill=X, expand=YES)
        path_lbl = ttk.Label(path_row, text="路径", width=8)
        path_lbl.pack(side=LEFT, padx=(15, 0))
        path_ent = ttk.Entry(path_row, textvariable=self.path_var)
        path_ent.pack(side=LEFT, fill=X, expand=YES, padx=5)
        windnd.hook_dropfiles(path_ent, func=self.dragged_files)
        browse_btn = ttk.Button(
            master=path_row,
            text="选择文件夹",
            command=self.on_browse,
            width=8
        )
        browse_btn.pack(side=LEFT, padx=5)

    def create_type_row(self):
        """Add type row to labelframe"""

        type_row = ttk.Frame(self.option_lf)
        type_row.pack(fill=X, expand=YES, pady=15)
        type_lbl = ttk.Label(type_row, text="模式", width=8)
        type_lbl.pack(side=LEFT, padx=(15, 0))

        contains_opt = ttk.Radiobutton(
            master=type_row,
            text="复制,统计",
            variable=self.type_var,
            value="复制,统计"
        )
        contains_opt.pack(side=LEFT, padx=25)

        startswith_opt = ttk.Radiobutton(
            master=type_row,
            text="直接,改名",
            variable=self.type_var,
            value="直接,改名"
        )
        startswith_opt.pack(side=LEFT, padx=30)

        contains_opt.invoke()

        search_btn = ttk.Button(
            master=type_row,
            text="执行开始",
            command=self.start_use,
            bootstyle=OUTLINE,
            width=8
        )
        search_btn.pack(side=LEFT, padx=5)

    def on_browse(self):
        """Callback for directory browse"""
        path = askdirectory(title="Browse directory")
        if path:
            self.path_var.set(path)

    def dragged_files(self, files):
        msg = '\n'.join((item.decode('gbk') for item in files))
        self.path_var.set(msg)

    def start_use(self):
        self.progressbar.start(10)
        # start_now = Thread(target=self.file_use)
        # start_now.start()



    def file_use(self):
        """Recursively search directory for matching files"""
        use_type = self.type_var.get()
        use_root_path = self.path_var.get()
        root_dir = (use_root_path.replace(
            "\\", '/').replace("//", '/')).strip()
        filepaths, filenames, folderpaths = FileSearchEngine.read(root_dir)
        if use_type == '复制,统计':
            FileSearchEngine.save_pdf(
                filepaths, filenames, folderpaths, (root_dir+"_导出"), root_dir)
        elif use_type == '直接,改名':
            FileSearchEngine.rename_pdf(filepaths, filenames)

    @staticmethod
    def re_text(bt, text):
        m1 = re.search(bt, text)
        if m1 is not None:
            return FileSearchEngine.re_block(m1[0])

    @staticmethod
    def re_block(text):
        return text.replace(' ', '').replace('　', '').replace('）', '').replace(')', '').replace('：', ':')

    @staticmethod
    def get_all(dir_path):
        all_files_2 = []
        all_names_2 = []
        folder_path = []

        for root, sub_dirs, file_names in os.walk(dir_path):
            all_files = []
            all_names = []
            root = root.replace('\\', '/')
            for name in file_names:
                all_path = os.path.join(root, name)
                all_path = all_path.replace('\\', '/')
                all_files.append(all_path)
                all_names.append(name)
            all_files_2.append(all_files)
            all_names_2.append(all_names)
            folder_path.append(root)
        return all_files_2, all_names_2, folder_path

    @staticmethod
    def read(root_dir):
        all_files_2, all_names_2, folder_paths = FileSearchEngine.get_all(
            root_dir)
        return all_files_2, all_names_2, folder_paths

    @staticmethod
    def new_folder(i, root_dir, out_dir, old_dir):
        new_dir = old_dir.replace(root_dir, out_dir)
        if i != 0:
            new_sheet = old_dir.replace(root_dir + "/", "")
            new_sheet = new_sheet.replace("/", "_")
        else:
            new_sheet = old_dir.split("/")[-1]
        if not os.path.exists(new_dir):
            # 如果目标路径不存在原文件夹的话就创建
            os.makedirs(new_dir)
        if os.path.exists(new_dir):
            # 如果目标路径存在原文件夹的话就先删除
            shutil.rmtree(new_dir)
            os.makedirs(new_dir)
        return new_dir, new_sheet

    @staticmethod
    def check_name(pdf_text, invoice_name=None, tax_id=None):

        if "91440300MA5GC316X2" in pdf_text or "9 1440300MA5GC316X2" in pdf_text:
            tax_id = "对了"
            if "佛山市顺德区瑞磐科技有限公司龙华分公司" in pdf_text:
                invoice_name = "对了"
            else:
                invoice_name = "×了"
        else:
            tax_id = "×了"
            if "佛山市顺德区瑞磐科技有限公司龙华分公司" in pdf_text:
                invoice_name = "对了"
            else:
                invoice_name = "×了"
        return invoice_name, tax_id

    @staticmethod
    def save_pdf(filepaths, filenames, folderpaths, out_dir, root_dir):
        for i, filepath in enumerate(filepaths):
            Repeat_name_list = []
            Repeat_num_list = []
            new_out_dir, new_sheet = FileSearchEngine.new_folder(
                i, root_dir, out_dir, folderpaths[i])
            if i == 0:
                new_root_dir = new_out_dir
            for j, filepathNeed in enumerate(filepath):
                if filenames[i][j].endswith('.pdf') or filenames[i][j].endswith('.PDF'):
                    tax_id = "没检查"
                    invoice_name = "没检查"
                    with pdfplumber.open(filepathNeed) as pdf:
                        first_page = pdf.pages[0]
                        pdf_text = first_page.extract_text()
                        invoice_name, tax_id = FileSearchEngine.check_name(
                            pdf_text, invoice_name, tax_id)
                        try:
                            list_excel = FileSearchEngine.re_info_1(
                                pdf_text, tax_id, invoice_name, filenames[i][j])
                            out_file_name = list_excel[4] + \
                                "-" + str(list_excel[5])
                        except:
                            list_excel = FileSearchEngine.re_info_2(
                                filenames[i][j])
                            out_file_name = list_excel[4]

                    list_excel, Repeat_name_list = FileSearchEngine.find_repeat_name(
                        list_excel, Repeat_name_list, out_file_name)

                    list_excel, Repeat_num_list = FileSearchEngine.find_repeat_num(
                        list_excel, Repeat_num_list)

                    FileSearchEngine.copy_rename(
                        i, filepathNeed, new_out_dir, new_root_dir, new_sheet, list_excel)
                else:
                    new_out = filepathNeed.replace(root_dir, new_root_dir)
                    shutil.copy(filepathNeed, new_out)
        # Messagebox.okcancel(title='提示', message="程序运行完了")

    @staticmethod
    def copy_rename(i, filepath, out_dir, new_root_dir, new_sheet, list_excel):
        dst = os.path.join(out_dir, list_excel[7])
        shutil.copy(filepath, dst)
        FileSearchEngine.save_excel_2(i, new_root_dir, sheet_name=new_sheet, value_list=list_excel)

    @staticmethod
    def find_repeat_name(list_excel, Repeat_name_list, out_file_name):
        Repeat_name_list.append(out_file_name)
        if out_file_name in Repeat_name_list:
            repeat_num = Repeat_name_list.count(out_file_name)
            if repeat_num == 1:
                out_file_name = out_file_name
            elif list_excel[3] == "重复":
                out_file_name = out_file_name + "(" + "重复" + ")"
            else:
                out_file_name = out_file_name + "(" + str(repeat_num-1) + ")"
            list_excel[7] = out_file_name + ".pdf"
        return list_excel, Repeat_name_list

    @staticmethod
    def find_repeat_num(list_excel, Repeat_num_list):
        if list_excel is not None:
            Repeat_num_list.append(list_excel[2])
            if list_excel[2] in Repeat_num_list:
                repeat_num = Repeat_num_list.count(list_excel[2])
                if repeat_num > 1:
                    list_excel[3] = "重复"
        return list_excel, Repeat_num_list

    @staticmethod
    def re_info_1(pdf_text, tax_id, invoice_name, filename):
        list_excel = []

        invoice_number = (FileSearchEngine.re_text(
            re.compile(r'发票号码(.*\d+)'), pdf_text))
        invoice_number = invoice_number.split(":", 1)[-1]

        invoice_date = (FileSearchEngine.re_text(
            re.compile(r'开票日期(.*)'), pdf_text))
        invoice_date = invoice_date.split(":", 1)[-1]

        invoice_type_true = (
            FileSearchEngine.re_text(re.compile(r'([/*]+[\u4e00-\u9fa5]+[ ])'), pdf_text))
        if invoice_type_true is None:
            invoice_type_true = (
                FileSearchEngine.re_text(re.compile(r'([/（]+[\u4e00-\u9fa5]+[/）])'), pdf_text))
            invoice_type_true = invoice_type_true.split("（", 1)[-1]
        invoice_type_true = invoice_type_true.split("*", 1)[-1]

        if "服务费" in pdf_text and "电费" in pdf_text and "住宿" not in pdf_text:
            invoice_type_true = "充电费"
        elif "服务费" in pdf_text and "住宿" in pdf_text:
            invoice_type_true = "住宿费"

        total_price_ture = FileSearchEngine.re_text(
            re.compile(r'(小写.*(.*[0-9.]+))'), pdf_text)
        if total_price_ture is None:
            total_price_ture = re.findall(
                (r'(￥.*[0-9.]+)|(¥.*[0-9.]+)'), pdf_text)[-1]
            total_price_ture = ''.join(total_price_ture)
        if "￥" in total_price_ture:
            total_price_ture = total_price_ture.split("￥", 1)[-1]
        elif "¥" in total_price_ture:
            total_price_ture = total_price_ture.split("¥", 1)[-1]

        list_excel.append(invoice_name)
        list_excel.append(tax_id)
        list_excel.append(int(invoice_number))
        list_excel.append("无重复")
        list_excel.append(invoice_type_true)
        list_excel.append(float(total_price_ture))
        list_excel.append(invoice_date)
        list_excel.append("")
        list_excel.append(filename)
        return list_excel

    @staticmethod
    def re_info_2(filename):
        list_excel = []

        list_excel.append("")
        list_excel.append("")
        list_excel.append("")
        list_excel.append("")
        list_excel.append("其他单或读不出来")
        list_excel.append("")
        list_excel.append("")
        list_excel.append("")
        list_excel.append(filename)
        return list_excel

    @staticmethod
    def save_excel_2(i, path, sheet_name, value_list=[[]]):
        value_list = [value_list]
        path = path + "/自动汇总统计.xlsx"
        try:
            workbook = openpyxl.load_workbook(path)
            Sheet = workbook[sheet_name]
        except:
            try:
                workbook = openpyxl.load_workbook(path)
            except:
                workbook = openpyxl.Workbook()
            workbook.create_sheet(sheet_name, i+1)
            Sheet = workbook[sheet_name]
            title = ['抬头', '纳税号', '发票号码', '发票',
                     '发票类型', '金额汇总', '开票日期', '改后pdf名称', '原pdf名称']
            Sheet.append(title)
            workbook.save(path)
        # it_exist = FileSearchEngine.find_exist(Sheet, value_list)
        # if not it_exist:
        if True:
            Sheet.append(value_list[0])
            for item, value in enumerate(value_list[0]):
                Sheet.column_dimensions[get_column_letter(
                    item+1)].width = 1.8*len(str(value)) + 3
            workbook.save(path)
        workbook.close()

    @staticmethod
    def find_exist(Sheet, value_list):
        for row in Sheet.rows:
            row_data = [cell.value if cell.value !=
                        None else "" for cell in row]
            if value_list[0] == row_data:
                return True
        return False

    @staticmethod
    def rename_pdf(filepaths, filenames):
        for i, filepath in enumerate(filepaths):
            Rename_list = []
            for j, filepathNeed in enumerate(filepath):
                if filenames[i][j].endswith('.pdf') or filenames[i][j].endswith('.PDF'):
                    try:
                        filenameNeed = filenames[i][j].split("-")[-1]
                        if "(" in filenameNeed:
                            filenameNeed = filenameNeed.split("(")[0] + ".pdf"
                        Rename_list.append(filenameNeed)
                        new_name = filepathNeed.replace(
                            filenames[i][j], filenameNeed)
                        if filenameNeed in Rename_list:
                            repeat_num = Rename_list.count(filenameNeed)
                            if repeat_num == 1:
                                new_name = new_name
                            else:
                                new_name = new_name.replace(".PDF", "").replace(
                                    ".pdf", "") + "(" + str(repeat_num) + ")" + ".pdf"
                        os.rename(filepathNeed, new_name)
                    except:
                        pass
        # print("okcancel: ",Messagebox.okcancel(title='提示', message="程序运行完了"))
        Messagebox.okcancel(title='提示', message="程序运行完了")


if __name__ == '__main__':

    app = ttk.Window(
        title="File Tools",  # 设置窗口的标题
        themename="journal",  # 设置主题
        # size=(1066,600),        #窗口的大小
        # position=(100,100),     #窗口所在的位置
        # minsize=(0,0),          #窗口的最小宽高
        # maxsize=(1920,1080),    #窗口的最大宽高
        # resizable=None,         #设置窗口是否可以更改大小
        # alpha=1.0,              #设置窗口的透明度(0.0完全透明）
    )

    FileSearchEngine(app)

    # app.place_window_center()  # 让显现出的窗口居中
    app.resizable(False, False)  # 让窗口不可更改大小
    # app.wm_attributes('-topmost', 1)#让窗口位置其它窗口之上
    
    app.mainloop()
    