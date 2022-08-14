import time
import openpyxl
from threading import Thread


def lcSheet(Workbook, sheet_name):
    try:
        Workbook = openpyxl.load_workbook(path)
    except:
        Workbook = openpyxl.Workbook()
    try:
        for name in sheet_name:
            Sheet = Workbook[name]
    except:
        for name in sheet_name:
            Workbook.create_sheet(name)
    Workbook.save(path)
    return Workbook


def rSheet(Workbook, sheet_name):
    Sheet = Workbook[sheet_name]
    return Sheet


def wSheet(Sheet, Workbook, value_list):
    value_list = ["1", "2", "3"]
    value_list = [value_list]
    Sheet.append(value_list[0])
    Workbook.save(path)
    return Workbook


def do_it(Workbook, name):
    # Workbook = lcSheet(path, sheet_name)
    Sheet = rSheet(Workbook, name)
    Workbook = wSheet(Sheet, Workbook, value_list=None)


def main(path, sheet_name):
    Workbook = lcSheet(path, sheet_name)
    # do_it(Workbook, sheet_name)
    for name in sheet_name:
        Thread(
            target=do_it,
            args=(Workbook, name),
            daemon=True
        ).start()
    Workbook.close()


if __name__ == '__main__':

    path = "D:/ProgramData/Design/demo/QWE/test.xlsx"

    sheet_name = ["asd", "qwe", "123", "456"]

    main(path, sheet_name)
